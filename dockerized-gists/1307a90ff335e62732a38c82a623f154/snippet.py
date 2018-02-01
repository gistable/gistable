# crawler.py
# Python 2.7.6


"""
Crawl a page and extract all urls recursively within same domain
"""


from BeautifulSoup import BeautifulSoup
from selenium import webdriver
from urlparse import urlparse
import pandas as pd
import openpyxl as op


parent_url = "https://automatetheboringstuff.com/"
domain = urlparse(parent_url).netloc

xl_name = "Crawler.xlsx"
sheet_name = "URLs"


def crawl_urls(url_list, crawled_urls, driver, url):
    """ get a set of urls and crawl each url recursively"""

    # Once the url is parsed, add it to crawled url list
    crawled_urls.append(url)

    driver.get(url)
    html = driver.page_source.encode("utf-8")

    soup = BeautifulSoup(html)

    urls = soup.findAll("a")

    # Even if the url is not part of the same domain, it is still collected
    # But those urls not in the same domain are not parsed
    for a in urls:
        if (a.get("href")) and (a.get("href") not in url_list):
            url_list.append(a.get("href"))

    # Recursively parse each url within same domain
    for page in set(url_list):  # set to remove duplicates

        # Check if the url belong to the same domain
        # And if this url is already parsed ignore it
        if (urlparse(page).netloc == domain) and (page not in crawled_urls):

            # print this_url
            crawl_urls(url_list, crawled_urls, driver, page)

    # Once all urls are crawled return the list to calling function
    else:
        return crawled_urls, url_list


def load_to_excel(lst):
    """
    Load the list into excel file using pandas
    """
    # Load list to dataframe
    df = pd.DataFrame(lst)
    df.index += 1  # So that the excel column starts from 1

    # Write dataframe to excel
    xlw = pd.ExcelWriter(xl_name)
    df.to_excel(xlw, sheet_name=sheet_name, index_label="#", header=["URL"])
    xlw.save()


def format_excel(xl, sheet="Sheet1"):
    """
    Get the excel file path and format the file
    If no sheet name is passed, by default take Sheet1
    """
    # Open the excel file
    wb = op.load_workbook(xl)
    ws = wb.get_sheet_by_name(sheet)

    # Freeze panes
    ws.freeze_panes = "B2"

    # Adjust column width
    cols = ("A", "B")
    widths = (5, 80)

    for combo in zip(cols, widths):
        ws.column_dimensions[combo[0]].width = combo[1]

    # define color formmatting
    blue_fill = op.styles.PatternFill(start_color="00aadd",
                                      fill_type='solid')

    # define border style
    thin_border = op.styles.borders.Border(left=op.styles.Side(style='thin'),
                                           right=op.styles.Side(style='thin'),
                                           top=op.styles.Side(style='thin'),
                                           bottom=op.styles.Side(style='thin'))

    # define Text wrap
    text_wrap = op.styles.Alignment(wrap_text=True)

    # Format the header row
    for row in range(1, 2):  # Loop only the 1st row
        for col in range(1, ws.max_column + 1):  # loop through all columns
            ws.cell(row=row, column=col).fill = blue_fill

    # Format all cells
    for row in ws.iter_rows():
        for cell in row:
            # Draw borders
            cell.border = thin_border

            # Wrap all columns
            cell.alignment = text_wrap

    # Save back as same file name
    wb.save(xl)


if __name__ == "__main__":
    """ Starting block """

    driver = webdriver.Firefox()

    url_list = list()
    crawled_urls = list()

    url_list.append(parent_url)

    # Initiate the crawling by passind the beginning url
    crawled_urls, url_list = crawl_urls(url_list, crawled_urls,
                                        driver, parent_url)

    # Finally quit the browser
    driver.quit()

    print "FULL URLs LIST"
    print len(set(url_list))

    print "CRAWLED URLs LIST"
    print len(set(crawled_urls))

    # Load the match list to excel
    load_to_excel(url_list)

    # Format the excel file
    format_excel(xl_name, sheet_name)
